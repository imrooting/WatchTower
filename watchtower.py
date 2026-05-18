#!/usr/bin/env python3
"""
WatchTower - Automated Reconnaissance & OSINT Framework
Modular, extensible, production-ready recon tooling.
"""

import sys
import os
import re
import subprocess
import argparse
import shutil
import glob
import time
import datetime
import socket
import getpass
from pathlib import Path

# ─────────────────────────────────────────────
# Ensure we're using Python 3.10+
# ─────────────────────────────────────────────
if sys.version_info < (3, 10):
    print("[-] WatchTower requires Python 3.10 or higher.")
    sys.exit(1)

# ─────────────────────────────────────────────
# Path setup
# ─────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent
MODULES_DIR = BASE_DIR / "modules"
sys.path.insert(0, str(BASE_DIR))

# ─────────────────────────────────────────────
# Dependency bootstrap
# ─────────────────────────────────────────────
def _ensure(package: str, import_name: str | None = None) -> None:
    import_name = import_name or package
    try:
        __import__(import_name)
    except ImportError:
        print(f"[~] Installing missing dependency: {package} …")
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet", package],
            check=True,
        )

_ensure("xlsxwriter")
_ensure("openpyxl")

try:
    import readline
    import rlcompleter  # noqa: F401
    _HAS_READLINE = True
except ImportError:
    _HAS_READLINE = False

# ─────────────────────────────────────────────
# Colour helpers
# ─────────────────────────────────────────────
_NO_COLOR = not sys.stdout.isatty() or bool(os.environ.get("NO_COLOR"))

def _c(code: str, text: str) -> str:
    return text if _NO_COLOR else f"\033[{code}m{text}\033[0m"

def red(t: str) -> str:    return _c("31", t)
def green(t: str) -> str:  return _c("32", t)
def yellow(t: str) -> str: return _c("33", t)
def blue(t: str) -> str:   return _c("34", t)
def cyan(t: str) -> str:   return _c("36", t)
def bold(t: str) -> str:   return _c("1",  t)

def info(msg: str) -> None:    print(blue(f"[*] {msg}"))
def success(msg: str) -> None: print(green(f"[+] {msg}"))
def warn(msg: str) -> None:    print(yellow(f"[!] {msg}"))
def error(msg: str) -> None:   print(red(f"[-] {msg}"), file=sys.stderr)

def progress(current: int, total: int, label: str = "") -> None:
    if total == 0:
        return
    pct = current / total * 100
    bar_len = 30
    filled = int(bar_len * current / total)
    bar = "█" * filled + "░" * (bar_len - filled)
    suffix = f" {label}" if label else ""
    print(f"\r  {cyan(bar)} {pct:5.1f}%{suffix}", end="", flush=True)
    if current == total:
        print()

# ─────────────────────────────────────────────
# ASCII banner
# ─────────────────────────────────────────────
BANNER = cyan(r"""
 ██╗    ██╗ █████╗ ████████╗ ██████╗██╗  ██╗████████╗ ██████╗ ██╗    ██╗███████╗██████╗
 ██║    ██║██╔══██╗╚══██╔══╝██╔════╝██║  ██║╚══██╔══╝██╔═══██╗██║    ██║██╔════╝██╔══██╗
 ██║ █╗ ██║███████║   ██║   ██║     ███████║   ██║   ██║   ██║██║ █╗ ██║█████╗  ██████╔╝
 ██║███╗██║██╔══██║   ██║   ██║     ██╔══██║   ██║   ██║   ██║██║███╗██║██╔══╝  ██╔══██╗
 ╚███╔███╔╝██║  ██║   ██║   ╚██████╗██║  ██║   ██║   ╚██████╔╝╚███╔███╔╝███████╗██║  ██║
  ╚══╝╚══╝ ╚═╝  ╚═╝   ╚═╝    ╚═════╝╚═╝  ╚═╝   ╚═╝    ╚═════╝  ╚══╝╚══╝ ╚══════╝╚═╝  ╚═╝
""") + bold("  Automated Reconnaissance & OSINT Framework  ") + blue("v1.0.0\n")

# ─────────────────────────────────────────────
# Module registry
# ─────────────────────────────────────────────
MODULE_REGISTRY: dict[str, dict] = {
    "nmap":         {"desc": "Comprehensive network scan via Nmap",              "fleet": True,  "ext": False},
    "spidering":    {"desc": "Crawl domains with hakrawler & gospider",          "fleet": True,  "ext": False},
    "sslscan":      {"desc": "Scan SSL/TLS for misconfigurations",               "fleet": True,  "ext": False},
    "secretfinder": {"desc": "Find secrets in JavaScript files",                 "fleet": True,  "ext": False},
    "subdomains":   {"desc": "Discover subdomains via assetfinder/amass",        "fleet": True,  "ext": False},
    "whatweb":      {"desc": "Identify web technologies",                        "fleet": True,  "ext": False},
    "waf":          {"desc": "Detect WAF protections",                           "fleet": True,  "ext": False},
    "waybackurls":  {"desc": "Pull historical URLs from Wayback Machine",        "fleet": True,  "ext": False},
    "httpx":        {"desc": "Enumerate live HTTP/HTTPS servers",                "fleet": True,  "ext": False},
    "nuclei":       {"desc": "Auto-detect vulns via Nuclei templates",           "fleet": True,  "ext": False},
    "dnsrecon":     {"desc": "DNS reconnaissance",                               "fleet": False, "ext": False},
    "getheaders":   {"desc": "Capture & save HTTP response headers",             "fleet": False, "ext": False},
    "getbody":      {"desc": "Capture & save HTTP response bodies",              "fleet": False, "ext": False},
    "spiderfoot":   {"desc": "Gather emails, names, phones via SpiderFoot",      "fleet": False, "ext": False},
    "formfinder":   {"desc": "Extract web forms via SpiderFoot",                 "fleet": False, "ext": False},
    "asn":          {"desc": "Fetch ASN, IP block, geolocation → Excel",         "fleet": False, "ext": False},
    "cve":          {"desc": "Search CVEs via Shodan/ASN → Excel",               "fleet": False, "ext": False},
    "github":       {"desc": "Find GitHub repos related to targets",             "fleet": False, "ext": False},
    "finddocs":     {"desc": "Extract URLs by file extension",                   "fleet": False, "ext": True },
    "gitleaks":     {"desc": "Scan git repos for secrets",                       "fleet": False, "ext": False},
    "ffuf":         {"desc": "Directory & endpoint bruteforcing via ffuf",        "fleet": True,  "ext": False},
}

MODULES      = list(MODULE_REGISTRY.keys())
AXIOM_MODULES = [k for k, v in MODULE_REGISTRY.items() if v["fleet"]]

# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT DIRECTORY STRUCTURE
# ─────────────────────────────────────────────────────────────────────────────
#
#  watchtower-results/
#  ├── MANIFEST.txt                  ← index of every produced file
#  ├── final/                        ← analyst-facing deliverables
#  │   ├── nmap/
#  │   ├── spidering/
#  │   ├── sslscan/
#  │   ├── subdomains/
#  │   ├── secretfinder/
#  │   ├── whatweb/
#  │   ├── waf/
#  │   ├── waybackurls/
#  │   ├── httpx/
#  │   ├── nuclei/
#  │   ├── dnsrecon/
#  │   ├── getheaders/
#  │   ├── getbody/
#  │   ├── spiderfoot/
#  │   ├── formfinder/
#  │   ├── asn/
#  │   ├── cve/
#  │   ├── github/
#  │   ├── finddocs/
#  │   └── gitleaks/
#  └── debug/                        ← raw/intermediate artefacts (auto-cleaned)
#      └── <module>/
#
# ─────────────────────────────────────────────────────────────────────────────

# Global session metadata (populated in main())
_SESSION: dict = {}


def _ts() -> str:
    """ISO-8601 timestamp for filenames (no colons)."""
    return datetime.datetime.now().strftime("%Y%m%dT%H%M%S")


def _now_human() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def init_session(args: argparse.Namespace) -> dict:
    """Populate global session metadata used in file headers."""
    global _SESSION
    _SESSION = {
        "module":      args.module,
        "input_file":  os.path.abspath(args.input),
        "fleet":       args.fleet or "N/A",
        "extension":   args.extension or "N/A",
        "operator":    getpass.getuser(),
        "hostname":    socket.gethostname(),
        "started_at":  _now_human(),
        "timestamp":   _ts(),
        "version":     "WatchTower v1.0.0",
        "targets":     read_lines(args.input) if os.path.isfile(args.input) else [],
    }
    return _SESSION


def _output_root() -> Path:
    return Path("watchtower-results")


def final_dir(module: str) -> Path:
    p = _output_root() / "final" / module
    p.mkdir(parents=True, exist_ok=True)
    return p


def debug_dir(module: str) -> Path:
    p = _output_root() / "debug" / module
    p.mkdir(parents=True, exist_ok=True)
    return p


def _write_text_header(module: str, description: str, extra_lines: list[str] | None = None) -> str:
    """
    Return a multi-line comment block to prepend to any plain-text output file.
    Includes what the file is, how it was generated, and from what input.
    """
    s = _SESSION
    targets_preview = "\n".join(f"#   {t}" for t in s.get("targets", [])[:20])
    if len(s.get("targets", [])) > 20:
        targets_preview += f"\n#   ... and {len(s['targets']) - 20} more"

    extra = ""
    if extra_lines:
        extra = "\n" + "\n".join(f"# {ln}" for ln in extra_lines)

    return (
        f"# {'═' * 76}\n"
        f"# WatchTower Reconnaissance Report\n"
        f"# {'─' * 76}\n"
        f"# Module      : {module}\n"
        f"# Description : {description}\n"
        f"# Generated   : {s.get('started_at', 'unknown')}\n"
        f"# Operator    : {s.get('operator', 'unknown')} @ {s.get('hostname', 'unknown')}\n"
        f"# Tool        : {s.get('version', 'WatchTower')}\n"
        f"# Input File  : {s.get('input_file', 'unknown')}\n"
        f"# Fleet       : {s.get('fleet', 'N/A')}\n"
        f"# Targets ({len(s.get('targets', []))}):\n"
        f"{targets_preview}\n"
        f"#{extra}\n"
        f"# {'═' * 76}\n\n"
    )


def _xlsx_add_meta_sheet(wb, module: str, description: str) -> None:
    """Add a 'Report Info' sheet as the first sheet in any Excel workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet("Report Info", 0)
    s  = _SESSION

    TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    LABEL_FONT  = Font(name="Calibri", bold=True, size=11)
    VALUE_FONT  = Font(name="Calibri", size=11)
    TITLE_FILL  = PatternFill("solid", fgColor="1F4E79")
    LABEL_FILL  = PatternFill("solid", fgColor="D6E4F0")
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    # Title row
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value     = "WatchTower Reconnaissance Report"
    title_cell.font      = TITLE_FONT
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    rows = [
        ("Module",       module),
        ("Description",  description),
        ("Generated",    s.get("started_at", "unknown")),
        ("Operator",     f"{s.get('operator', 'unknown')} @ {s.get('hostname', 'unknown')}"),
        ("Tool Version", s.get("version", "WatchTower")),
        ("Input File",   s.get("input_file", "unknown")),
        ("Fleet",        s.get("fleet", "N/A")),
        ("Target Count", str(len(s.get("targets", [])))),
        ("Targets",      "\n".join(s.get("targets", []))),
    ]

    for r, (label, value) in enumerate(rows, 2):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font      = LABEL_FONT
        lc.fill      = LABEL_FILL
        lc.alignment = LEFT

        vc = ws.cell(row=r, column=2, value=value)
        vc.font      = VALUE_FONT
        vc.alignment = LEFT

    # Tall row for targets
    ws.row_dimensions[len(rows) + 1].height = max(15, len(s.get("targets", [])) * 15)


def write_manifest_entry(module: str, filepath: str, description: str) -> None:
    """Append a line to the global MANIFEST.txt."""
    manifest = _output_root() / "MANIFEST.txt"
    _output_root().mkdir(parents=True, exist_ok=True)
    with open(manifest, "a", encoding="utf-8") as f:
        rel = os.path.relpath(filepath, _output_root())
        f.write(f"[{_now_human()}] [{module}] {rel:<60} | {description}\n")


def _init_manifest() -> None:
    """Write the manifest header at session start."""
    _output_root().mkdir(parents=True, exist_ok=True)
    manifest = _output_root() / "MANIFEST.txt"
    s = _SESSION
    with open(manifest, "a", encoding="utf-8") as f:
        f.write(
            f"\n{'═' * 100}\n"
            f"  WatchTower Session  |  Module: {s['module']}  |  Started: {s['started_at']}\n"
            f"  Operator: {s['operator']} @ {s['hostname']}  |  Input: {s['input_file']}\n"
            f"{'═' * 100}\n"
            f"  {'Timestamp':<22}  {'Module':<14}  {'File':<60}  Description\n"
            f"{'─' * 100}\n"
        )


def clean_debug(module: str) -> None:
    """Remove the debug sub-directory for a module after final outputs are written."""
    d = debug_dir(module)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
        info(f"Cleaned debug artefacts for module '{module}'")


# ─────────────────────────────────────────────
# Tab completion
# ─────────────────────────────────────────────
def _setup_completion() -> None:
    if not _HAS_READLINE:
        return
    candidates = MODULES + ["--module", "--input", "--fleet", "--extension",
                             "--output-dir", "--verbose", "--help",
                             "-m", "-i", "-f", "-e", "-o", "-v"]
    def completer(text: str, state: int) -> str | None:
        options = [c for c in candidates if c.startswith(text)]
        return options[state] if state < len(options) else None
    readline.set_completer(completer)
    readline.parse_and_bind("tab: complete")

# ─────────────────────────────────────────────
# Core utilities
# ─────────────────────────────────────────────
def require_file(path: str, label: str = "Input file") -> None:
    if not path or not os.path.isfile(path):
        error(f"{label} '{path}' not found.")
        sys.exit(1)

def require_fleet(fleet: str, module: str) -> None:
    if not fleet:
        error(f"Module '{module}' requires --fleet / -f <fleet_name>")
        sys.exit(1)

def require_extension(ext: str, module: str) -> None:
    if not ext:
        error(f"Module '{module}' requires --extension / -e <ext>")
        sys.exit(1)

def require_module_script(name: str) -> Path:
    path = MODULES_DIR / name
    if not path.is_file():
        error(f"Module helper '{name}' not found in {MODULES_DIR}")
        sys.exit(1)
    return path

def run_cmd(cmd: str, *, check: bool = True,
            verbose: bool = False) -> subprocess.CompletedProcess:
    if verbose:
        info(f"Running: {cmd}")
    result = subprocess.run(cmd, shell=True)
    if check and result.returncode != 0:
        error(f"Command failed (exit {result.returncode}): {cmd}")
        sys.exit(result.returncode)
    return result

def axiom_scan(input_file: str, fleet: str, module_flag: str,
               output: str, extra: str = "",
               verbose: bool = False) -> None:
    cmd = (f'axiom-scan "{input_file}" --fleet "{fleet}" '
           f'-m {module_flag} -o "{output}"')
    if extra:
        cmd += f" {extra}"
    run_cmd(cmd, verbose=verbose)

def read_lines(path: str) -> list[str]:
    with open(path, encoding="utf-8", errors="replace") as f:
        return [ln.strip() for ln in f if ln.strip()]

def mkdir(path: str | Path) -> None:
    os.makedirs(path, exist_ok=True)

def move_to_debug(files_or_patterns: list[str], module: str) -> None:
    """Move intermediate files into the debug directory for this module."""
    dest = debug_dir(module)
    for pat in files_or_patterns:
        for f in glob.glob(pat):
            try:
                shutil.move(f, dest)
            except shutil.Error:
                pass

def copy_to_final(src: str | Path, module: str,
                  description: str = "") -> Path:
    """
    Copy a finished artefact into final/<module>/ and register it
    in the manifest. Returns the destination path.
    """
    dest_dir  = final_dir(module)
    dest_path = dest_dir / Path(src).name
    shutil.copy2(src, dest_path)
    write_manifest_entry(module, str(dest_path), description or Path(src).name)
    return dest_path

def ask(prompt: str, choices: list[str]) -> str:
    print()
    for i, c in enumerate(choices, 1):
        print(f"  {cyan(str(i))}) {c}")
    while True:
        raw = input(f"\n{prompt} [1-{len(choices)}]: ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(choices):
            return choices[int(raw) - 1]
        warn("Invalid choice — try again.")

def yes_no(prompt: str) -> bool:
    ans = input(f"{blue(prompt)} [y/N]: ").strip().lower()
    return ans in ("y", "yes")

def run_module_script(script_name: str, *args: str,
                      check: bool = True) -> subprocess.CompletedProcess:
    script = require_module_script(script_name)
    cmd = [sys.executable, str(script)] + list(args)
    info(f"Running helper: {script_name} {' '.join(args)}")
    result = subprocess.run(cmd)
    if check and result.returncode != 0:
        error(f"Helper '{script_name}' failed (exit {result.returncode})")
        sys.exit(result.returncode)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def _excel_styles():
    """Return a dict of commonly used openpyxl styles."""
    from openpyxl.styles import Font, PatternFill, Alignment
    return {
        "header_font":  Font(bold=True, color="FFFFFF", name="Calibri"),
        "header_fill":  PatternFill("solid", fgColor="1F4E79"),
        "alt_fill":     PatternFill("solid", fgColor="D6E4F0"),
        "white_fill":   PatternFill("solid", fgColor="FFFFFF"),
        "red_fill":     PatternFill("solid", fgColor="FFDEDE"),
        "center":       Alignment(horizontal="center", vertical="center", wrap_text=True),
        "top_wrap":     Alignment(vertical="top", wrap_text=True),
        "top":          Alignment(vertical="top"),
    }


def _write_header_row(ws, headers: list[str], styles: dict) -> None:
    from openpyxl.utils import get_column_letter
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = styles["header_font"]
        c.fill      = styles["header_fill"]
        c.alignment = styles["center"]


def _set_col_widths(ws, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze_and_filter(ws) -> None:
    ws.freeze_panes     = "A2"
    ws.auto_filter.ref  = ws.dimensions


# ─────────────────────────────────────────────────────────────────────────────
# MODULE IMPLEMENTATIONS
# ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════
# nmap
# ══════════════════════════════════════════════
def run_nmap(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "nmap"
    require_file(input_file)
    require_fleet(fleet, MODULE)
    require_module_script("convert-nmap-to-xml.py")

    DEFAULT_TCP = (
        "-p 20,21,22,23,69,3389,5900-5902,512-514,873,53,111,2049,"
        "135,137,138,139,445,161,389,25,110,143,80,8000,8080,8888,"
        "8443,1433,1521,3306,5000,5432,6379,27017-27018"
    )
    DEFAULT_UDP = "53,67,68,69,161,162,123,514,520,137,138,139,5355,5353"

    scan_type = ask("Select scan type", ["TCP Scan", "UDP Scan"])
    tcp_opt = udp_opt = tcp_ports = udp_ports = defeat_icmp = ""

    if scan_type == "TCP Scan":
        tcp_type = ask("Select TCP scan type",
                       ["TCP Connect Scan (-sT)", "SYN Scan (-sS)"])
        tcp_opt = "-sT" if "sT" in tcp_type else "-sS"
        port_choice = ask("Select TCP ports",
                          ["Full port scan (-p-)", "Common ports (default)", "Custom ports"])
        if port_choice.startswith("Full"):
            tcp_ports = "-p-"
        elif port_choice.startswith("Common"):
            tcp_ports = DEFAULT_TCP
        else:
            custom = input("  Enter custom ports (e.g. 80,443): ").strip()
            tcp_ports = f"-p {custom}"
    else:
        udp_opt = "-sU"
        port_choice = ask("Select UDP ports",
                          ["Full port scan (-p-)", "Common UDP ports (default)", "Custom ports"])
        if port_choice.startswith("Full"):
            udp_ports = "-p-"
        elif port_choice.startswith("Common"):
            udp_ports = f"-p {DEFAULT_UDP}"
        else:
            custom = input("  Enter custom ports (e.g. 53,161): ").strip()
            udp_ports = f"-p {custom}"
        defeat_icmp = "--defeat-icmp-ratelimit"

    speed       = ask("Select scan speed", ["T3 (Normal)", "T4 (Aggressive)", "T5 (Insane)"])
    speed_flag  = f"-T{speed[1]}"
    scripts_flag = "--script=default" if yes_no("Run default scripts?") else ""
    version_flag = "-sV" if yes_no("Enumerate service versions?") else ""

    # All intermediate nmap artefacts land in debug/nmap/
    dbg = debug_dir(MODULE)

    args_str = " ".join(filter(None, [
        tcp_ports, udp_ports, "--open", speed_flag,
        scripts_flag, version_flag, tcp_opt, udp_opt, defeat_icmp,
        '--script-args "http.useragent=Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko"',
        "--script-timeout 60m", "--host-timeout 60m", "-oA nmapx",
    ]))

    cmd = f'axiom-scan "{input_file}" --fleet "{fleet}" -m nmapx {args_str}'
    success(f"Nmap command:\n  {yellow(cmd)}")
    run_cmd(cmd, verbose=verbose)

    # Merge raw axiom XML output
    info("Merging Nmap XML files …")
    run_module_script("convert-nmap-to-xml.py")

    merged_xml = "merged_nmap_output.xml"
    if not os.path.isfile(merged_xml):
        error("merged_nmap_output.xml not found after conversion.")
        sys.exit(1)

    # Move raw/intermediate files to debug
    move_to_debug(
        ["nmapx*", "scan+*", merged_xml],
        MODULE,
    )
    merged_xml_in_debug = str(dbg / "merged_nmap_output.xml")

    # Parse → structured Excel in final/nmap/
    out_xlsx = str(final_dir(MODULE) / f"nmap-results-{_SESSION['timestamp']}.xlsx")
    info(f"Generating Excel report → {out_xlsx}")
    _nmap_xml_to_xlsx(merged_xml_in_debug, out_xlsx)

    # Also write a plain-text summary of open ports
    out_txt = str(final_dir(MODULE) / f"nmap-open-ports-{_SESSION['timestamp']}.txt")
    _nmap_xml_to_txt(merged_xml_in_debug, out_txt)

    write_manifest_entry(MODULE, out_xlsx, "Nmap structured Excel report (3 sheets: Open Ports, Host Summary, Vulnerabilities)")
    write_manifest_entry(MODULE, out_txt,  "Nmap plain-text open-port summary")

    success(f"Nmap complete. Final outputs → {final_dir(MODULE)}/")


def _nmap_xml_to_xlsx(xml_path: str, out_path: str) -> None:
    import xml.etree.ElementTree as ET
    from openpyxl import Workbook

    if not os.path.isfile(xml_path):
        error(f"XML file '{xml_path}' not found.")
        sys.exit(1)
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except ET.ParseError as exc:
        error(f"Failed to parse '{xml_path}': {exc}")
        sys.exit(1)

    wb  = Workbook()
    st  = _excel_styles()

    # ── Sheet 0: Report Info ──────────────────────────────────────────────
    _xlsx_add_meta_sheet(wb, "nmap", "Comprehensive network scan via Nmap")

    # ── Sheet 1: Open Ports ───────────────────────────────────────────────
    ws_ports = wb.create_sheet("Open Ports")
    PORT_HDRS = ["IP Address", "Hostname", "Port", "Protocol",
                 "State", "Service", "Version", "Script Output"]
    _write_header_row(ws_ports, PORT_HDRS, st)

    row_idx = 2
    for host in root.findall("host"):
        ip = next(
            (a.get("addr", "") for a in host.findall("address") if a.get("addrtype") == "ipv4"),
            ""
        )
        hostnames = host.find("hostnames")
        hostname  = ""
        if hostnames is not None:
            hn = hostnames.find("hostname")
            if hn is not None:
                hostname = hn.get("name", "")

        ports_el = host.find("ports")
        if ports_el is None:
            continue

        for port in ports_el.findall("port"):
            state_el = port.find("state")
            if state_el is None or state_el.get("state") != "open":
                continue

            portid   = port.get("portid", "")
            protocol = port.get("protocol", "")
            state    = state_el.get("state", "")

            service_el = port.find("service")
            service = version = ""
            if service_el is not None:
                service = service_el.get("name", "")
                version = " ".join(filter(None, [
                    service_el.get("product", ""),
                    service_el.get("version", ""),
                    service_el.get("extrainfo", ""),
                ])).strip()

            scripts = [
                f"[{s.get('id','')}] {s.get('output','').strip()}"
                for s in port.findall("script") if s.get("output", "").strip()
            ]
            script_out = "\n".join(scripts)

            fill = st["alt_fill"] if row_idx % 2 == 0 else st["white_fill"]
            for col, val in enumerate([ip, hostname, portid, protocol,
                                        state, service, version, script_out], 1):
                c = ws_ports.cell(row=row_idx, column=col, value=val)
                c.fill      = fill
                c.alignment = st["top_wrap"]
            row_idx += 1

    _set_col_widths(ws_ports, [16, 28, 8, 10, 10, 16, 32, 60])
    _freeze_and_filter(ws_ports)

    # ── Sheet 2: Host Summary ─────────────────────────────────────────────
    ws_hosts = wb.create_sheet("Host Summary")
    HOST_HDRS = ["IP Address", "Hostname", "OS Match", "OS Accuracy %",
                 "Status", "Open Port Count"]
    _write_header_row(ws_hosts, HOST_HDRS, st)

    row_idx = 2
    for host in root.findall("host"):
        ip = next(
            (a.get("addr", "") for a in host.findall("address") if a.get("addrtype") == "ipv4"),
            ""
        )
        hostnames = host.find("hostnames")
        hostname  = ""
        if hostnames is not None:
            hn = hostnames.find("hostname")
            if hn is not None:
                hostname = hn.get("name", "")

        status_el = host.find("status")
        status    = status_el.get("state", "") if status_el is not None else ""

        os_match = os_acc = ""
        os_el    = host.find("os")
        if os_el is not None:
            best = os_el.find("osmatch")
            if best is not None:
                os_match = best.get("name", "")
                os_acc   = best.get("accuracy", "")

        ports_el   = host.find("ports")
        open_count = 0
        if ports_el is not None:
            open_count = sum(
                1 for p in ports_el.findall("port")
                if (s := p.find("state")) is not None and s.get("state") == "open"
            )

        fill = st["alt_fill"] if row_idx % 2 == 0 else st["white_fill"]
        for col, val in enumerate([ip, hostname, os_match, os_acc,
                                    status, open_count], 1):
            c = ws_hosts.cell(row=row_idx, column=col, value=val)
            c.fill      = fill
            c.alignment = st["top"]
        row_idx += 1

    _set_col_widths(ws_hosts, [16, 28, 40, 14, 10, 16])
    _freeze_and_filter(ws_hosts)

    # ── Sheet 3: Vulnerabilities ──────────────────────────────────────────
    ws_vuln = wb.create_sheet("Vulnerabilities")
    VULN_HDRS = ["IP Address", "Port", "Script ID", "Finding"]
    _write_header_row(ws_vuln, VULN_HDRS, st)

    VULN_RE = re.compile(
        r"VULNERABLE|CVE-|CWE-|exploit|overflow|injection|disclosure|bypass|RCE", re.I
    )
    row_idx = 2
    for host in root.findall("host"):
        ip = next(
            (a.get("addr", "") for a in host.findall("address") if a.get("addrtype") == "ipv4"),
            ""
        )
        ports_el = host.find("ports")
        if ports_el is None:
            continue
        for port in ports_el.findall("port"):
            portid = port.get("portid", "")
            for script in port.findall("script"):
                sid    = script.get("id", "")
                output = script.get("output", "").strip()
                if output and VULN_RE.search(output):
                    for col, val in enumerate([ip, portid, sid, output], 1):
                        c = ws_vuln.cell(row=row_idx, column=col, value=val)
                        c.fill      = st["red_fill"]
                        c.alignment = st["top_wrap"]
                    row_idx += 1

    _set_col_widths(ws_vuln, [16, 8, 28, 80])
    _freeze_and_filter(ws_vuln)

    wb.save(out_path)


def _nmap_xml_to_txt(xml_path: str, out_path: str) -> None:
    """Write a clean plain-text open-port summary with a metadata header."""
    import xml.etree.ElementTree as ET

    header = _write_text_header(
        "nmap",
        "Open ports extracted from Nmap XML output",
        ["Format: IP | Hostname | Port/Protocol | Service | Version"],
    )

    lines: list[str] = []
    if os.path.isfile(xml_path):
        try:
            root = ET.parse(xml_path).getroot()
            for host in root.findall("host"):
                ip = next(
                    (a.get("addr", "") for a in host.findall("address")
                     if a.get("addrtype") == "ipv4"), ""
                )
                hostnames = host.find("hostnames")
                hostname  = ""
                if hostnames is not None:
                    hn = hostnames.find("hostname")
                    if hn is not None:
                        hostname = hn.get("name", "")
                ports_el = host.find("ports")
                if ports_el is None:
                    continue
                for port in ports_el.findall("port"):
                    state_el = port.find("state")
                    if state_el is None or state_el.get("state") != "open":
                        continue
                    portid   = port.get("portid", "")
                    protocol = port.get("protocol", "")
                    svc_el   = port.find("service")
                    service  = svc_el.get("name", "") if svc_el is not None else ""
                    version  = ""
                    if svc_el is not None:
                        version = " ".join(filter(None, [
                            svc_el.get("product", ""),
                            svc_el.get("version", ""),
                        ]))
                    lines.append(
                        f"{ip:<18} {hostname:<30} {portid}/{protocol:<8} "
                        f"{service:<16} {version}"
                    )
        except ET.ParseError:
            lines.append("# ERROR: Could not parse XML.")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(header)
        f.write("\n".join(lines) + "\n")


# ══════════════════════════════════════════════
# getheaders
# ══════════════════════════════════════════════
def run_getheaders(input_file: str, verbose: bool = False) -> None:
    MODULE = "getheaders"
    require_file(input_file)
    require_module_script("capture-headers.py")

    dbg  = debug_dir(MODULE)
    fin  = final_dir(MODULE)
    urls = read_lines(input_file)
    if not urls:
        error("Input file is empty.")
        sys.exit(1)

    raw_dir = dbg / "raw-headers"
    raw_dir.mkdir(parents=True, exist_ok=True)

    info(f"Fetching headers for {len(urls)} URL(s) …")
    for i, url in enumerate(urls, 1):
        safe = re.sub(r"https?://", "", url).replace("/", "_")
        safe = re.sub(r"[^\w._-]", "", safe)[:200]
        out  = raw_dir / (safe + ".txt")
        result = subprocess.run(
            ["curl", "-L", "-D", str(out), "-k", "--max-time", "30", url, "-o", "/dev/null"],
            capture_output=True,
        )
        if verbose and result.returncode != 0:
            warn(f"curl failed for {url} (exit {result.returncode})")
        progress(i, len(urls), url)

    # Merge raw header files
    merged_raw = dbg / "allresponseheaders-raw.txt"
    with open(merged_raw, "w", encoding="utf-8") as mf:
        for f in sorted(raw_dir.iterdir()):
            mf.write(f"Filename: {f.name}\n")
            mf.write(f.read_text(errors="replace"))
            mf.write("\n")

    # Generate Excel (helper writes response-headers-output.xlsx in cwd)
    info("Converting to Excel …")
    run_module_script("capture-headers.py", str(merged_raw))

    tmp_xlsx = Path("response-headers-output.xlsx")
    if tmp_xlsx.exists():
        # Inject metadata sheet before moving
        _inject_meta_into_xlsx(tmp_xlsx, MODULE, "HTTP response headers captured via curl")
        ts_name  = f"response-headers-{_SESSION['timestamp']}.xlsx"
        dest_xlsx = fin / ts_name
        shutil.move(str(tmp_xlsx), str(dest_xlsx))
        write_manifest_entry(MODULE, str(dest_xlsx), "HTTP response headers — all targets (Excel)")

    # Also write a plain-text merged copy with header block
    out_txt = fin / f"response-headers-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(MODULE, "HTTP response headers captured via curl -D")
    raw_content = merged_raw.read_text(errors="replace")
    with open(out_txt, "w", encoding="utf-8") as f:
        f.write(header)
        f.write(raw_content)
    write_manifest_entry(MODULE, str(out_txt), "HTTP response headers — plain text merged")

    success(f"getheaders complete. Final outputs → {fin}/")


def _inject_meta_into_xlsx(path: Path, module: str, description: str) -> None:
    """Open an existing xlsx and prepend a Report Info sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(path)
    _xlsx_add_meta_sheet(wb, module, description)
    wb.save(path)


# ══════════════════════════════════════════════
# getbody
# ══════════════════════════════════════════════
def run_getbody(input_file: str, verbose: bool = False) -> None:
    MODULE = "getbody"
    require_file(input_file)

    dbg  = debug_dir(MODULE)
    fin  = final_dir(MODULE)
    urls = read_lines(input_file)
    if not urls:
        error("Input file is empty.")
        sys.exit(1)

    raw_dir = dbg / "raw-bodies"
    raw_dir.mkdir(parents=True, exist_ok=True)

    info(f"Fetching response bodies for {len(urls)} URL(s) …")
    for i, url in enumerate(urls, 1):
        safe = re.sub(r"https?://", "", url).replace("/", "_")
        safe = re.sub(r"[^\w._-]", "", safe)[:200]
        out  = raw_dir / (safe + ".txt")
        subprocess.run(
            ["curl", "-o", str(out), "-k", "--max-time", "30", url],
            capture_output=not verbose,
        )
        progress(i, len(urls), url)

    # Merge + annotate
    out_txt = fin / f"response-bodies-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(MODULE, "HTTP response bodies captured via curl")

    with open(out_txt, "w", encoding="utf-8") as mf:
        mf.write(header)
        for f in sorted(raw_dir.iterdir()):
            mf.write(f"\n{'─' * 60}\n")
            mf.write(f"# Source URL file : {f.name}\n")
            mf.write(f"{'─' * 60}\n")
            mf.write(f.read_text(errors="replace"))

    write_manifest_entry(MODULE, str(out_txt), "HTTP response bodies — all targets merged")
    success(f"getbody complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# spidering
# ══════════════════════════════════════════════
def run_spidering(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "spidering"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    lines = read_lines(input_file)
    if not any("://" in ln for ln in lines):
        new_file = str(dbg / "https_input_file.txt")
        warn(f"No scheme detected; writing https:// prefixed copy → {new_file}")
        with open(new_file, "w") as f:
            f.writelines(f"https://{ln}\n" for ln in lines)
        input_file = new_file

    if not shutil.which("jq"):
        info("jq not found — installing …")
        run_cmd("sudo apt-get install -y jq", verbose=verbose)

    axiom_scan(input_file, fleet, "hakrawler", str(dbg / "hakrawler-raw.txt"), verbose=verbose)
    run_cmd(f'cat "{dbg}/hakrawler-raw.txt" | jq -r \'.URL\' > "{dbg}/hakrawler-urls.txt"', verbose=verbose)
    run_cmd(f'grep -Ff "{input_file}" "{dbg}/hakrawler-urls.txt" > "{dbg}/hakrawler-filtered.txt" || true', verbose=verbose)

    axiom_scan(input_file, fleet, "gospider", str(dbg / "gospider-output"), verbose=verbose)
    run_cmd(f'cat "{dbg}/gospider-output/"* | awk \'{{print $3}}\' | grep http | grep -v \'^$\' > "{dbg}/gospider-urls.txt"', verbose=verbose)
    run_cmd(f'grep -Ff "{input_file}" "{dbg}/gospider-urls.txt" > "{dbg}/gospider-filtered.txt" || true', verbose=verbose)

    combined = str(dbg / "combined-urls.txt")
    run_cmd(f'cat "{dbg}/hakrawler-filtered.txt" "{dbg}/gospider-filtered.txt" | sort -u > "{combined}"', verbose=verbose)

    _create_final_url_list(combined, fleet, fin, dbg, verbose)
    _create_final_js_url_list(combined, fleet, fin, dbg, verbose)

    # Annotate final URL list
    for fname, desc in [
        ("Final-URL-List.txt",            "Deduplicated live URLs (non-JS)"),
        ("Final-JavaScript-URL-List.txt", "Live JavaScript/CSS asset URLs"),
    ]:
        src = fin / fname
        if src.exists():
            _prepend_text_header(src, MODULE, desc)
            write_manifest_entry(MODULE, str(src), desc)

    success(f"Spidering complete. Final outputs → {fin}/")


def _prepend_text_header(path: Path, module: str, description: str) -> None:
    header  = _write_text_header(module, description)
    content = path.read_text(errors="replace")
    path.write_text(header + content, encoding="utf-8")


def _create_final_url_list(combined: str, fleet: str,
                            fin: Path, dbg: Path, verbose: bool) -> None:
    MODULE = "spidering"
    t1 = str(dbg / "url-list-temp1.txt")
    t2 = str(dbg / "url-list-temp2.txt")
    t3 = str(dbg / "url-list-temp3.txt")

    run_cmd(f'cat "{combined}" | grep -Ev \'\\.js$|\\.css$|jquery|css|/js/|js\' > "{t1}"', verbose=verbose)
    axiom_scan(t1, fleet, "httpx", t2, verbose=verbose)
    run_cmd(f'cat "{t2}" | grep -v -E \'404|FAILED|400|401|403\' | awk \'{{print $1}}\' > "{t3}"', verbose=verbose)

    if not shutil.which("qsreplace"):
        info("Installing qsreplace …")
        run_cmd("go install github.com/tomnomnom/qsreplace@latest", verbose=verbose)
        os.environ["PATH"] += ":" + os.path.expanduser("~/go/bin")

    run_cmd(f'cat "{t3}" | qsreplace -a > "{fin}/Final-URL-List.txt"', verbose=verbose)


def _create_final_js_url_list(combined: str, fleet: str,
                               fin: Path, dbg: Path, verbose: bool) -> None:
    t1 = str(dbg / "js-list-temp1.txt")
    t2 = str(dbg / "js-list-temp2.txt")
    run_cmd(f'cat "{combined}" | grep -i -E \'jquery|js|css\' > "{t1}"', verbose=verbose)
    axiom_scan(t1, fleet, "httpx", t2, verbose=verbose)
    run_cmd(f'cat "{t2}" | grep -v -E \'404|FAILED|400|401|403\' | awk \'{{print $1}}\' > "{fin}/Final-JavaScript-URL-List.txt"', verbose=verbose)


# ══════════════════════════════════════════════
# secretfinder
# ══════════════════════════════════════════════
def run_secretfinder(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "secretfinder"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    warn("Output may contain false positives — review carefully.")
    time.sleep(2)

    info(f"Ensuring SecretFinder is installed on fleet '{fleet}' …")
    run_cmd(
        "axiom-exec '"
        "if [ ! -f ~/recon/SecretFinder/SecretFinder.py ]; then "
        "  echo \"[fleet] SecretFinder not found — cloning …\" && "
        "  mkdir -p ~/recon && cd ~/recon && "
        "  git clone --depth=1 https://github.com/m4ll0k/SecretFinder.git && "
        "  cd SecretFinder && pip install -r requirements.txt --quiet && "
        "  echo \"[fleet] SecretFinder installed.\"; "
        "else "
        "  echo \"[fleet] SecretFinder already present.\"; "
        "fi'",
        verbose=verbose,
    )

    raw_out = str(dbg / "secretfinder-raw.txt")
    run_cmd(f'axiom-scan "{input_file}" --fleet "{fleet}" -m secretfinder | tee -a "{raw_out}"', verbose=verbose)

    if not os.path.isfile(raw_out):
        error("secretfinder raw output not created.")
        sys.exit(1)

    with open(raw_out, encoding="utf-8", errors="replace") as f:
        content = f.read()

    lines, in_block = [], False
    for line in content.splitlines():
        if "[ + ]" in line:
            in_block = True
        if in_block:
            lines.append(line)

    filtered = [
        ln for ln in lines
        if not re.search(r"home|fire|module|runtime|mode|URL: \$url", ln, re.I)
    ]

    out_txt = fin / f"javascript-secrets-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(
        MODULE,
        "Secrets found in JavaScript files via SecretFinder",
        ["WARNING: Output may contain false positives. Review before actioning."],
    )
    with open(out_txt, "w") as f:
        f.write(header)
        f.write("\n".join(filtered))

    write_manifest_entry(MODULE, str(out_txt), "JavaScript secrets (filtered SecretFinder output)")
    success(f"secretfinder complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# sslscan
# ══════════════════════════════════════════════
def run_sslscan(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "sslscan"
    require_file(input_file)
    require_fleet(fleet, MODULE)
    require_module_script("pretty_parse_ssl_alt_names.py")

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    info(f"Ensuring sslscan is installed on fleet '{fleet}' …")
    run_cmd(
        "axiom-exec '"
        "if ! command -v sslscan &>/dev/null; then "
        "  echo \"[fleet] sslscan not found — building from source …\" && "
        "  git clone --depth=1 https://github.com/rbsec/sslscan.git /tmp/sslscan-build && "
        "  cd /tmp/sslscan-build && make static && "
        "  sudo cp sslscan /usr/local/bin/sslscan && "
        "  rm -rf /tmp/sslscan-build && "
        "  echo \"[fleet] sslscan installed: $(sslscan --version)\"; "
        "else "
        "  echo \"[fleet] sslscan already present: $(sslscan --version)\"; "
        "fi'",
        verbose=verbose,
    )

    raw_out = str(dbg / "sslscan-raw.txt")
    axiom_scan(input_file, fleet, "sslscan", raw_out, verbose=verbose)

    if not os.path.isfile(raw_out):
        error("sslscan-raw.txt not created.")
        sys.exit(1)

    ansi_escape = re.compile(r"\x1B\[[0-9;]*[a-zA-Z]")
    raw         = open(raw_out, encoding="utf-8", errors="replace").read()
    clean       = ansi_escape.sub("", raw)

    clean_file = dbg / "sslscan-clean.txt"
    clean_file.write_text(clean, encoding="utf-8")

    if not clean_file.stat().st_size:
        warn("sslscan output is empty — nothing to parse.")
        return

    # ── SSL/TLS protocol status ───────────────────────────────────────────
    status_lines: list[str] = []
    for ln in clean.splitlines():
        if re.search(r"SSL|TLS", ln, re.I) and not re.search(
            r"Accepted|Preferred|SSL Certificate:|TLS Compression|TLS renegotiation"
            r"|TLS Fallback|heartbleed|SL Certificate:", ln, re.I
        ):
            parts = ln.split()
            if len(parts) >= 11 and not re.search(
                r"openssl|128|192|224|260|RSA|SSL/TLS Protocols:", ln, re.I
            ):
                status_lines.append(f"{parts[0]} {parts[1]} {parts[10]}")

    ssl_status_file = fin / f"ssl-tls-status-{_SESSION['timestamp']}.txt"
    header = _write_text_header(MODULE, "SSL/TLS protocol status per host")
    with open(ssl_status_file, "w") as f:
        f.write(header)
        f.write("\n".join(status_lines))
    write_manifest_entry(MODULE, str(ssl_status_file), "SSL/TLS protocol enable/disable status")

    # ── Weak ciphers ──────────────────────────────────────────────────────
    ip = port = ""
    weak: list[str] = []
    for ln in clean.splitlines():
        m = re.match(r"Testing SSL server (.+?) on port (\d+)", ln)
        if m:
            ip, port = m.group(1), m.group(2)
            continue
        if re.search(r"null|rc4|rc2|des|sm4|maga|cnt|md5|sm3", ln, re.I):
            weak.append(f"{ip}:{port} {ln.strip()}")

    weak_file = fin / f"weak-ciphers-{_SESSION['timestamp']}.txt"
    header    = _write_text_header(MODULE, "Weak/deprecated SSL ciphers detected",
                                   ["These ciphers should be disabled on the target servers."])
    with open(weak_file, "w") as f:
        f.write(header)
        f.write("\n".join(weak))
    write_manifest_entry(MODULE, str(weak_file), "Weak/deprecated SSL ciphers")

    # ── Alt names / domains from certs ───────────────────────────────────
    run_module_script("pretty_parse_ssl_alt_names.py")

    alt_raw = open("ssl-alt-names.txt", encoding="utf-8", errors="replace").read() \
              if os.path.isfile("ssl-alt-names.txt") else ""
    shutil.move("ssl-alt-names.txt", str(dbg / "ssl-alt-names.txt")) \
        if os.path.isfile("ssl-alt-names.txt") else None

    domains: set[str] = set()
    for m in re.finditer(r"Altnames: (.+)", alt_raw):
        for entry in m.group(1).split(", "):
            d = entry.strip().lstrip("DNS:").strip()
            if d and not d.startswith("*"):
                domains.add(d)

    domain_file = fin / f"domains-from-ssl-cert-{_SESSION['timestamp']}.txt"
    header      = _write_text_header(MODULE, "Domain names extracted from SSL certificate alt-names")
    with open(domain_file, "w") as f:
        f.write(header)
        f.write("\n".join(sorted(domains)))
    write_manifest_entry(MODULE, str(domain_file), "Domains extracted from SSL certificate SANs")

    # Move any leftover scan artefacts to debug
    move_to_debug(["scan+*", "ssl-tls-status-temp.txt"], MODULE)

    success(f"sslscan complete. Final outputs → {fin}/")


# ══════════════════════════════════════════════
# subdomains
# ══════════════════════════════════════════════
def run_subdomains(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "subdomains"
    require_file(input_file)
    require_fleet(fleet, MODULE)
    require_module_script("parse_domains_subdomains.py")

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    scan = ask("Select scan depth",
               ["Quick Scan (Assetfinder only)", "Deep Scan (Assetfinder + Amass)"])

    af_out   = str(dbg / "assetfinder_output.txt")
    httpx_in = af_out

    if "Quick" in scan:
        axiom_scan(input_file, fleet, "assetfinder", af_out, verbose=verbose)
    else:
        am_out    = str(dbg / "amass_output.txt")
        merged    = str(dbg / "merged_unique_subdomains.txt")
        axiom_scan(input_file, fleet, "assetfinder", af_out, verbose=verbose)
        axiom_scan(input_file, fleet, "amass",       am_out, verbose=verbose)
        run_cmd(f'cat "{af_out}" "{am_out}" | sort -u > "{merged}"', verbose=verbose)
        httpx_in = merged

    httpx_out = str(dbg / "subdomain-httpx-result.txt")
    axiom_scan(httpx_in, fleet, "httpx", httpx_out, verbose=verbose)

    # Parse and write final outputs
    live_subs = _parse_live_subdomains(httpx_out)
    domains   = read_lines(input_file)

    # Plain-text live subdomains list
    live_file = fin / f"live-subdomains-{_SESSION['timestamp']}.txt"
    header    = _write_text_header(MODULE, "Live subdomains confirmed via httpx")
    with open(live_file, "w") as f:
        f.write(header)
        f.write("\n".join(live_subs))
    write_manifest_entry(MODULE, str(live_file), "Live subdomains (httpx confirmed)")

    # Grouped by parent domain
    grouped_file = fin / f"subdomains-by-domain-{_SESSION['timestamp']}.txt"
    header       = _write_text_header(MODULE, "Live subdomains grouped by parent domain")
    with open(grouped_file, "w") as f:
        f.write(header)
        for domain in domains:
            subs_for = [s for s in live_subs if s.endswith(f".{domain}")]
            f.write(f"\nDomain: {domain}  ({len(subs_for)} subdomains)\n")
            f.write("─" * 50 + "\n")
            for sub in subs_for:
                f.write(f"  {sub}\n")
    write_manifest_entry(MODULE, str(grouped_file), "Subdomains grouped by parent domain")

    # Excel summary
    xlsx_file = fin / f"subdomains-{_SESSION['timestamp']}.xlsx"
    _subdomains_to_xlsx(domains, live_subs, str(xlsx_file))
    write_manifest_entry(MODULE, str(xlsx_file), "Subdomains Excel report")

    # Run the bundled parser helper (writes its own files; move to debug)
    run_module_script("parse_domains_subdomains.py")
    move_to_debug(["subdomains_list.txt", "live_subdomains.txt", "temp_subdomain.txt"], MODULE)

    success(f"Subdomains complete. Final outputs → {fin}/")


def _parse_live_subdomains(httpx_result: str) -> list[str]:
    if not os.path.isfile(httpx_result):
        return []
    lines = read_lines(httpx_result)
    cleaned: list[str] = []
    for ln in lines:
        ln = re.sub(r"https?://(www\.)?", "", ln.split()[0]) if ln else ln
        if ln:
            cleaned.append(ln)
    return cleaned


def _subdomains_to_xlsx(domains: list[str], live_subs: list[str], out_path: str) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    st = _excel_styles()
    _xlsx_add_meta_sheet(wb, "subdomains", "Subdomain enumeration via assetfinder / amass + httpx")

    ws = wb.create_sheet("Live Subdomains")
    _write_header_row(ws, ["Parent Domain", "Subdomain"], st)

    row_idx = 2
    for domain in domains:
        for sub in [s for s in live_subs if s.endswith(f".{domain}")]:
            fill = st["alt_fill"] if row_idx % 2 == 0 else st["white_fill"]
            for col, val in enumerate([domain, sub], 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.fill      = fill
                c.alignment = st["top"]
            row_idx += 1

    _set_col_widths(ws, [28, 50])
    _freeze_and_filter(ws)
    wb.save(out_path)


# ══════════════════════════════════════════════
# whatweb
# ══════════════════════════════════════════════
def run_whatweb(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "whatweb"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    info(f"Ensuring whatweb is installed on fleet '{fleet}' …")
    run_cmd(
        "axiom-exec '"
        "if ! command -v whatweb &>/dev/null; then "
        "  echo \"[fleet] whatweb not found — installing via apt …\" && "
        "  sudo apt-get install -y whatweb && "
        "  echo \"[fleet] whatweb installed: $(whatweb --version 2>&1 | head -1)\"; "
        "else "
        "  echo \"[fleet] whatweb already present: $(whatweb --version 2>&1 | head -1)\"; "
        "fi'",
        verbose=verbose,
    )

    raw_out = str(dbg / "whatweb-raw.txt")
    axiom_scan(input_file, fleet, "whatweb", raw_out, verbose=verbose)

    if not os.path.isfile(raw_out):
        error("whatweb raw output not created.")
        sys.exit(1)

    ansi    = re.compile(r"\x1B\[([0-9]{1,2}(;[0-9]{1,2})?)?[mK]")
    raw     = open(raw_out, encoding="utf-8", errors="replace").read()
    no_ansi = ansi.sub("", raw)
    formatted = re.sub(r"\]([^\]]*?)$", r"]\1\n", no_ansi, flags=re.MULTILINE)

    out_txt = fin / f"whatweb-results-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(MODULE, "Web technology fingerprinting via WhatWeb")
    with open(out_txt, "w") as f:
        f.write(header)
        f.write(formatted)

    write_manifest_entry(MODULE, str(out_txt), "WhatWeb technology fingerprint results")
    success(f"whatweb complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# waf
# ══════════════════════════════════════════════
def run_waf(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "waf"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    raw_out = str(dbg / "waf-raw.txt")
    axiom_scan(input_file, fleet, "wafw00f", raw_out, verbose=verbose)

    if not os.path.isfile(raw_out):
        error("waf raw output not created.")
        sys.exit(1)

    ansi = re.compile(r"\x1b\[[0-9;]*m")
    lines: list[str] = []
    with open(raw_out, encoding="utf-8", errors="replace") as f:
        for line in f:
            if "site" in line.lower():
                lines.append(ansi.sub("", line))

    out_txt = fin / f"waf-results-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(MODULE, "WAF detection results via wafw00f")
    with open(out_txt, "w") as f:
        f.write(header)
        f.writelines(lines)

    write_manifest_entry(MODULE, str(out_txt), "WAF detection results")
    success(f"waf complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# spiderfoot
# ══════════════════════════════════════════════
def run_spiderfoot(input_file: str, verbose: bool = False) -> None:
    MODULE = "spiderfoot"
    require_file(input_file)

    fin  = final_dir(MODULE)
    dbg  = debug_dir(MODULE)
    urls = read_lines(input_file)
    if not urls:
        error("Input file is empty.")
        sys.exit(1)

    per_target_dir = dbg / "per-target"
    per_target_dir.mkdir(parents=True, exist_ok=True)

    info(f"Running SpiderFoot against {len(urls)} target(s) …")
    for i, url in enumerate(urls, 1):
        safe = re.sub(r"[^\w._-]", "", url)[:200]
        out  = per_target_dir / f"{safe}.tsv"
        with open(out, "a") as f:
            f.write(f"URL: {url}\n")
        run_cmd(
            f'spiderfoot -m sfp_spider,sfp_email,sfp_names,sfp_phone '
            f'-s "{url}" -q -F EMAILADDR,HUMAN_NAME,PHONE_NUMBER >> "{out}"',
            check=False, verbose=verbose,
        )
        progress(i, len(urls))

    merged_tsv = dbg / "merged-raw.tsv"
    run_cmd(f'cat "{per_target_dir}"/*.tsv | grep -Ev \'\\-e\' > "{merged_tsv}"', check=False)

    out_txt = fin / f"spiderfoot-osint-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(
        MODULE,
        "OSINT data (emails, names, phone numbers) gathered via SpiderFoot",
    )
    content = merged_tsv.read_text(errors="replace") if merged_tsv.exists() else ""
    with open(out_txt, "w") as f:
        f.write(header)
        f.write(content)

    write_manifest_entry(MODULE, str(out_txt), "OSINT: emails, names, phone numbers")
    success(f"spiderfoot complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# formfinder
# ══════════════════════════════════════════════
def run_formfinder(input_file: str, verbose: bool = False) -> None:
    MODULE = "formfinder"
    require_file(input_file)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    per_target = dbg / "per-target"
    per_target.mkdir(parents=True, exist_ok=True)

    for url in read_lines(input_file):
        safe = re.sub(r"[^\w._-]", "", url)[:200]
        out  = per_target / f"{safe}.tsv"
        with open(out, "a") as f:
            f.write(f"URL: {url}\n")
        run_cmd(
            f'spiderfoot -m sfp_pageinfo -s "{url}" -q -F URL_FORM >> "{out}"',
            check=False, verbose=verbose,
        )

    merged_tsv = dbg / "merged-forms.tsv"
    run_cmd(f'cat "{per_target}"/*.tsv > "{merged_tsv}"', check=False)

    out_txt = fin / f"forms-found-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(MODULE, "Web forms discovered via SpiderFoot sfp_pageinfo")
    content = merged_tsv.read_text(errors="replace") if merged_tsv.exists() else ""
    with open(out_txt, "w") as f:
        f.write(header)
        f.write(content)

    write_manifest_entry(MODULE, str(out_txt), "Web forms found across all targets")
    success(f"formfinder complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# dnsrecon
# ══════════════════════════════════════════════
def run_dnsrecon(input_file: str, verbose: bool = False) -> None:
    MODULE = "dnsrecon"
    require_file(input_file)
    require_module_script("dnsrecon-merge.py")

    if not shutil.which("dnsrecon"):
        error("dnsrecon not found. Install with: sudo apt install dnsrecon")
        sys.exit(1)

    fin     = final_dir(MODULE)
    dbg     = debug_dir(MODULE)
    domains = read_lines(input_file)
    if not domains:
        error("Input file is empty.")
        sys.exit(1)

    info(f"Running dnsrecon for {len(domains)} domain(s) …")
    for i, domain in enumerate(domains, 1):
        csv_out = str(dbg / f"{domain}_rec.csv")
        run_cmd(f'dnsrecon -d "{domain}" -t std -c "{csv_out}"',
                check=False, verbose=verbose)
        progress(i, len(domains), domain)

    # Merge helper writes merged_output.csv in cwd; move to debug first
    run_module_script("dnsrecon-merge.py")
    merged = "merged_output.csv"
    if os.path.isfile(merged):
        shutil.move(merged, str(dbg / "merged_output.csv"))

    merged_debug = str(dbg / "merged_output.csv")

    for grep_args, out_name, desc in [
        ("grep -vi txt",            f"dnsrecon-all-records-{_SESSION['timestamp']}.csv",   "All DNS records (no TXT)"),
        ("grep -iE 'v=DMARC|p='",  f"dmarc-records-{_SESSION['timestamp']}.csv",          "DMARC records"),
        ("grep -i 'v=spf'",         f"spf-records-{_SESSION['timestamp']}.csv",            "SPF records"),
    ]:
        out_path = fin / out_name
        run_cmd(f'{grep_args} "{merged_debug}" > "{out_path}"', check=False, verbose=verbose)
        # Prepend header block
        _prepend_text_header(out_path, MODULE, desc)
        write_manifest_entry(MODULE, str(out_path), desc)

    success(f"dnsrecon complete. Final outputs → {fin}/")


# ══════════════════════════════════════════════
# waybackurls
# ══════════════════════════════════════════════
def run_waybackurls(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "waybackurls"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin    = final_dir(MODULE)
    raw    = str(debug_dir(MODULE) / "waybackurls-raw.txt")
    axiom_scan(input_file, fleet, "waybackurls", raw, verbose=verbose)

    out_txt = fin / f"waybackurls-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(MODULE, "Historical URLs from the Wayback Machine (web.archive.org)")
    content = open(raw, encoding="utf-8", errors="replace").read() if os.path.isfile(raw) else ""
    with open(out_txt, "w") as f:
        f.write(header)
        f.write(content)

    write_manifest_entry(MODULE, str(out_txt), "Wayback Machine historical URLs")
    success(f"waybackurls complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# httpx
# ══════════════════════════════════════════════
def run_httpx(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "httpx"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin = final_dir(MODULE)
    raw = str(debug_dir(MODULE) / "httpx-raw.txt")
    axiom_scan(input_file, fleet, "httpx", raw, verbose=verbose)

    out_txt = fin / f"httpx-live-hosts-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(MODULE, "Live HTTP/HTTPS hosts enumerated via httpx")
    content = open(raw, encoding="utf-8", errors="replace").read() if os.path.isfile(raw) else ""
    with open(out_txt, "w") as f:
        f.write(header)
        f.write(content)

    write_manifest_entry(MODULE, str(out_txt), "Live HTTP/HTTPS hosts")
    success(f"httpx complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# nuclei
# ══════════════════════════════════════════════
def run_nuclei(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "nuclei"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin = final_dir(MODULE)
    raw = str(debug_dir(MODULE) / "nuclei-raw.txt")
    axiom_scan(input_file, fleet, "nuclei", raw, verbose=verbose)

    out_txt = fin / f"nuclei-findings-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(
        MODULE,
        "Vulnerability findings from Nuclei template scans",
        ["Severity: critical, high, medium, low, info"],
    )
    content = open(raw, encoding="utf-8", errors="replace").read() if os.path.isfile(raw) else ""
    with open(out_txt, "w") as f:
        f.write(header)
        f.write(content)

    write_manifest_entry(MODULE, str(out_txt), "Nuclei vulnerability scan findings")
    success(f"nuclei complete. Final output → {out_txt}")


# ══════════════════════════════════════════════
# asn
# ══════════════════════════════════════════════
def run_asn(input_file: str, verbose: bool = False) -> None:
    MODULE = "asn"
    require_file(input_file)
    require_module_script("capture-osint.py")

    if not shutil.which("asn"):
        info("Installing asn …")
        run_cmd(
            'sudo sh -c \'curl "https://raw.githubusercontent.com/nitefood/asn/master/asn" '
            '> /usr/bin/asn && chmod 0755 /usr/bin/asn\'',
            verbose=verbose,
        )

    fin = final_dir(MODULE)

    run_module_script("capture-osint.py", input_file)

    tmp = Path("asn.xlsx")
    if tmp.exists():
        _inject_meta_into_xlsx(tmp, MODULE, "ASN, IP block, and geolocation data")
        dest = fin / f"asn-results-{_SESSION['timestamp']}.xlsx"
        shutil.move(str(tmp), str(dest))
        write_manifest_entry(MODULE, str(dest), "ASN, IP blocks, geolocation (Excel)")

    success(f"asn complete. Final output → {fin}/")


# ══════════════════════════════════════════════
# cve
# ══════════════════════════════════════════════
def run_cve(input_file: str, verbose: bool = False) -> None:
    MODULE = "cve"
    require_file(input_file)
    import xlsxwriter as xlw
    from openpyxl import Workbook

    if not shutil.which("asn"):
        info("Installing asn …")
        run_cmd(
            'sudo sh -c \'curl "https://raw.githubusercontent.com/nitefood/asn/master/asn" '
            '> /usr/bin/asn && chmod 0755 /usr/bin/asn\'',
            verbose=verbose,
        )

    domains = read_lines(input_file)
    if not domains:
        error("Input file is empty.")
        sys.exit(1)

    fin = final_dir(MODULE)

    # Build with openpyxl so we can inject the meta sheet
    wb = Workbook()
    st = _excel_styles()
    _xlsx_add_meta_sheet(wb, MODULE, "CVE/vulnerability data from Shodan via ASN tool")

    ws = wb.create_sheet("CVE Results")
    _write_header_row(ws, ["Target", "Vulnerabilities"], st)

    info(f"Fetching CVEs for {len(domains)} domain(s) …")
    for i, domain in enumerate(domains, 1):
        cmd = (
            f"asn -J -n {domain} | jq -r "
            "'.results[].fingerprinting.vulns // [\"No vulnerabilities\"] | @csv' "
            "| sed 's/\"//g'"
        )
        try:
            out = subprocess.check_output(cmd, shell=True, timeout=60).decode().strip()
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
            out = "Error retrieving data"

        fill = st["alt_fill"] if i % 2 == 0 else st["white_fill"]
        for col, val in enumerate([domain, out], 1):
            c = ws.cell(row=i + 1, column=col, value=val)
            c.fill      = fill
            c.alignment = st["top_wrap"]
        progress(i, len(domains), domain)

    _set_col_widths(ws, [30, 80])
    _freeze_and_filter(ws)

    out_path = fin / f"cve-results-{_SESSION['timestamp']}.xlsx"
    wb.save(str(out_path))
    write_manifest_entry(MODULE, str(out_path), "CVE data per target (Excel)")

    success(f"cve complete. Final output → {out_path}")


# ══════════════════════════════════════════════
# github
# ══════════════════════════════════════════════
def run_github(input_file: str, verbose: bool = False) -> None:
    MODULE = "github"
    require_file(input_file)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    per_target = dbg / "per-target"
    per_target.mkdir(parents=True, exist_ok=True)

    for url in read_lines(input_file):
        safe = re.sub(r"[^\w._-]", "", url)[:200]
        out  = per_target / f"{safe}.tsv"
        with open(out, "a") as f:
            f.write(f"URL: {url}\n")
        run_cmd(
            f'spiderfoot -m sfp_spider,sfp_github -s "{url}" -q -F PUBLIC_CODE_REPO >> "{out}"',
            check=False, verbose=verbose,
        )

    merged    = dbg / "merged-github-raw.txt"
    run_cmd(f'cat "{per_target}"/*.tsv > "{merged}"', check=False)

    sorted_links = dbg / "github-links-sorted.txt"
    run_cmd(
        f'cat "{merged}" | grep -ivE \'\\-e|sfp_github|source|description\' '
        f'| sed \'s/URL://g\' | grep http | sed \'s/^URL: //\' | sort -u > "{sorted_links}"',
        check=False,
    )

    out_txt = fin / f"github-repos-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(
        MODULE,
        "GitHub repositories discovered via SpiderFoot",
        ["These links can be fed into the 'gitleaks' module for secret scanning."],
    )
    content = sorted_links.read_text(errors="replace") if sorted_links.exists() else ""
    with open(out_txt, "w") as f:
        f.write(header)
        f.write(content)

    # Also write gitleaks-ready copy (no header, just URLs)
    gitleaks_feed = fin / f"github-repos-for-gitleaks-{_SESSION['timestamp']}.txt"
    gitleaks_feed.write_text(content, encoding="utf-8")

    write_manifest_entry(MODULE, str(out_txt),       "GitHub repositories (annotated)")
    write_manifest_entry(MODULE, str(gitleaks_feed), "GitHub repo URLs — ready for gitleaks module")

    success(f"github complete. Final outputs → {fin}/")


# ══════════════════════════════════════════════
# finddocs
# ══════════════════════════════════════════════
def run_finddocs(input_file: str, extension: str, verbose: bool = False) -> None:
    MODULE = "finddocs"
    require_file(input_file)
    require_extension(extension, MODULE)
    from urllib.parse import urlparse

    fin  = final_dir(MODULE)
    urls = [ln for ln in read_lines(input_file) if ln.endswith(f".{extension}")]

    if not urls:
        warn(f"No URLs found with extension .{extension}")
        return

    # URL list
    out_urls = fin / f"{extension}-file-urls-{_SESSION['timestamp']}.txt"
    header   = _write_text_header(
        MODULE,
        f"URLs with .{extension} extension extracted from input",
        [f"Total found: {len(urls)}"],
    )
    with open(out_urls, "w") as f:
        f.write(header)
        f.write("\n".join(urls))
    write_manifest_entry(MODULE, str(out_urls), f"All .{extension} URLs")

    # Grouped by domain
    grouped: dict[str, list[str]] = {}
    for u in urls:
        domain = urlparse(u).netloc.lstrip("www.")
        grouped.setdefault(domain, []).append(u)

    out_grouped = fin / f"{extension}-files-by-domain-{_SESSION['timestamp']}.txt"
    header2     = _write_text_header(MODULE, f".{extension} URLs grouped by domain")
    with open(out_grouped, "w") as f:
        f.write(header2)
        for domain, links in sorted(grouped.items()):
            f.write(f"\nDomain: {domain}  ({len(links)} files)\n")
            f.write("─" * 50 + "\n")
            for link in links:
                f.write(f"  {link}\n")
    write_manifest_entry(MODULE, str(out_grouped), f".{extension} URLs grouped by domain")

    success(f"finddocs complete. {len(urls)} .{extension} URL(s) → {fin}/")


# ══════════════════════════════════════════════
# gitleaks
# ══════════════════════════════════════════════
def run_gitleaks(input_file: str, verbose: bool = False) -> None:
    MODULE = "gitleaks"
    require_file(input_file)
    require_module_script("clone-repos.py")

    if not shutil.which("gitleaks"):
        error("gitleaks not found. Install from https://github.com/gitleaks/gitleaks/releases")
        sys.exit(1)

    lines = read_lines(input_file)
    if not any("github.com" in ln for ln in lines):
        error("Input file must contain GitHub URLs (e.g. https://github.com/user/repo).")
        sys.exit(1)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    run_module_script("clone-repos.py", input_file)
    success("Repositories cloned.")

    cloned = Path("cloned_repos")
    if not cloned.is_dir():
        error("cloned_repos directory not found after cloning.")
        sys.exit(1)

    repos = [d for d in cloned.iterdir() if d.is_dir()]
    info(f"Running gitleaks on {len(repos)} repo(s) …")
    for i, repo in enumerate(repos, 1):
        info(f"  Scanning: {repo.name}")
        subprocess.run(
            ["gitleaks", "detect", "--report-format=json",
             "--report-path=gitleaks_report.json"],
            cwd=repo,
            capture_output=not verbose,
        )
        progress(i, len(repos), repo.name)

    # Collect reports with findings
    reports_dir = fin / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    found = 0
    for report in cloned.rglob("gitleaks_report.json"):
        if report.stat().st_size > 3:
            dest = reports_dir / f"{report.parent.name}_gitleaks_report.json"
            shutil.copy(report, dest)
            write_manifest_entry(MODULE, str(dest), f"Gitleaks findings: {report.parent.name}")
            found += 1

    # Move cloned repos to debug (they can be large; analyst may want to delete)
    shutil.move(str(cloned), str(dbg / "cloned_repos"))

    # Write a summary file
    summary = fin / f"gitleaks-summary-{_SESSION['timestamp']}.txt"
    header  = _write_text_header(
        MODULE,
        "Git repository secret scanning results via gitleaks",
        [f"Repos scanned : {len(repos)}",
         f"Repos with findings : {found}",
         "Individual JSON reports are in the reports/ sub-directory."],
    )
    with open(summary, "w") as f:
        f.write(header)
        f.write(f"Repos scanned      : {len(repos)}\n")
        f.write(f"Repos with findings: {found}\n\n")
        f.write("Reports:\n")
        for r in sorted(reports_dir.iterdir()):
            f.write(f"  {r.name}\n")
    write_manifest_entry(MODULE, str(summary), "Gitleaks scan summary")

    if found:
        success(f"Gitleaks found findings in {found} repo(s). Reports → {reports_dir}/")
    else:
        info("No gitleaks findings detected.")

    success(f"gitleaks complete. Final outputs → {fin}/")


# ══════════════════════════════════════════════
# ffuf  (directory & endpoint bruteforcing)
# ══════════════════════════════════════════════

def _get_fleet_nodes(fleet: str) -> list[str]:
    """Return sorted list of node names for the given fleet, e.g. ['pratik01', 'pratik02']."""
    import json as _json

    # Method 1: axiom-ls --json
    try:
        out = subprocess.check_output(
            ["axiom-ls", "--json"], timeout=15, stderr=subprocess.DEVNULL
        ).decode()
        data  = _json.loads(out)
        nodes = sorted(d["name"] for d in data if d.get("name", "").startswith(fleet))
        if nodes:
            return nodes
    except Exception:
        pass

    # Method 2: axiom-ls plain text
    try:
        out   = subprocess.check_output(["axiom-ls"], timeout=15, stderr=subprocess.DEVNULL).decode()
        nodes = sorted(
            line.split()[0] for line in out.splitlines()
            if line.strip().startswith(fleet)
        )
        if nodes:
            return nodes
    except Exception:
        pass

    # Method 3: ~/.axiom/selected.conf
    try:
        conf = Path.home() / ".axiom" / "selected.conf"
        if conf.is_file():
            nodes = sorted(
                ln.strip() for ln in conf.read_text().splitlines()
                if ln.strip().startswith(fleet)
            )
            if nodes:
                return nodes
    except Exception:
        pass

    return []


def run_ffuf(input_file: str, fleet: str, verbose: bool = False) -> None:
    MODULE = "ffuf"
    require_file(input_file)
    require_fleet(fleet, MODULE)

    fin = final_dir(MODULE)
    dbg = debug_dir(MODULE)

    # ── Ensure ffuf is on every fleet node ───────────────────────────────
    info(f"Ensuring ffuf is installed on fleet '{fleet}' …")
    run_cmd(
        f"axiom-exec 'command -v ffuf || "
        f"(go install github.com/ffuf/ffuf/v2@latest && "
        f"echo \"[fleet] ffuf installed\")' --fleet '{fleet}'",
        verbose=verbose,
    )

    # ── Wordlist selection ────────────────────────────────────────────────
    WORDLISTS = {
        "Quick  — common.txt (4k words, fastest)":
            "/usr/share/wordlists/dirb/common.txt",
        "Medium — directory-list-2.3-medium.txt (220k words)":
            "/usr/share/wordlists/dirbuster/directory-list-2.3-medium.txt",
        "Large  — directory-list-2.3-big.txt (1.2M words)":
            "/usr/share/wordlists/dirbuster/directory-list-2.3-big.txt",
        "Custom — enter local path":
            "__custom__",
    }
    wl_choice      = ask("Select wordlist", list(WORDLISTS.keys()))
    wordlist_local = WORDLISTS[wl_choice]
    upload_wordlist = False

    if wordlist_local == "__custom__":
        wordlist_local = input("  Enter full LOCAL path to wordlist: ").strip()
        require_file(wordlist_local, "Wordlist")
        upload_wordlist = True
    elif not os.path.isfile(wordlist_local):
        upload_wordlist = False   # system path already on fleet nodes
    else:
        upload_wordlist = True

    wl_name        = Path(wordlist_local).name
    wordlist_remote = f'"$HOME/recon/{wl_name}"' if upload_wordlist else f'"{wordlist_local}"'

    # ── Scan mode ─────────────────────────────────────────────────────────
    mode = ask(
        "Select scan mode",
        [
            "Directory bruteforce  (append FUZZ to URL)",
            "Subdomain bruteforce  (FUZZ.target.com)",
            "Parameter bruteforce  (url?FUZZ=value)",
            "Virtual host fuzzing  (Host: FUZZ.target.com header)",
        ],
    )

    ext_flag = ""
    if "Directory" in mode:
        if yes_no("Also fuzz file extensions (e.g. .php .html .txt)?"):
            exts     = input("  Enter extensions separated by commas (e.g. php,html,txt): ").strip()
            ext_flag = "-e " + ",".join("." + e.strip().lstrip(".") for e in exts.split(","))

    threads      = input("  Threads per node [default 40]: ").strip() or "40"
    timeout      = input("  Request timeout seconds [default 10]: ").strip() or "10"
    rate_limit   = input("  Max requests/sec per node (0 = unlimited) [default 0]: ").strip() or "0"
    filter_codes = input("  Filter OUT these HTTP status codes [default 404,400,500]: ").strip() or "404,400,500"

    extra_headers = ""
    if yes_no("Add custom HTTP headers?"):
        raw_hdrs = input("  Enter headers, semicolon-separated (e.g. Authorization: Bearer tok; X-Foo: bar): ").strip()
        for hdr in raw_hdrs.split(";"):
            hdr = hdr.strip()
            if hdr:
                extra_headers += f' -H "{hdr}"'

    follow_redirects = "-r" if yes_no("Follow redirects?") else ""
    rate_flag        = f"-rate {rate_limit}" if rate_limit != "0" else ""

    # ── Read targets ──────────────────────────────────────────────────────
    urls = read_lines(input_file)
    if not urls:
        error("Input file is empty.")
        sys.exit(1)

    # ── Discover fleet nodes and distribute targets ───────────────────────
    info(f"Discovering nodes in fleet '{fleet}' …")
    nodes = _get_fleet_nodes(fleet)

    if not nodes:
        warn(f"Could not enumerate nodes for fleet '{fleet}' — falling back to single-script mode.")
        nodes = [fleet]   # treat as one logical node

    success(f"Found {len(nodes)} node(s): {', '.join(nodes)}")
    info(f"Distributing {len(urls)} target(s) across {len(nodes)} node(s) …")

    # Round-robin split: target i goes to node i % len(nodes)
    node_targets: dict[str, list[str]] = {node: [] for node in nodes}
    for i, url in enumerate(urls):
        node_targets[nodes[i % len(nodes)]].append(url)

    for node, targets in node_targets.items():
        info(f"  {node} → {len(targets)} target(s): {', '.join(targets)}")

    # ── Generate one script per node ──────────────────────────────────────
    raw_out = dbg / "ffuf-raw-results"
    raw_out.mkdir(parents=True, exist_ok=True)

    per_node_scripts: dict[str, Path] = {}
    expected_jsons:   list[str]       = []

    for node, targets in node_targets.items():
        script_path = dbg / f"ffuf-targets-{node}.sh"
        per_node_scripts[node] = script_path

        with open(script_path, "w") as sh:
            sh.write("#!/usr/bin/env bash\n")
            sh.write(f"# WatchTower ffuf script for node: {node}\n")
            sh.write(f"# Targets ({len(targets)}): {', '.join(targets)}\n\n")
            sh.write('mkdir -p "$HOME/recon"\n\n')

            for url in targets:
                url         = url.strip().rstrip("/")
                bare        = re.sub(r"^https?://", "", url, flags=re.I)
                bare_no_www = re.sub(r"^www\.", "", bare, flags=re.I)
                canonical   = f"https://{bare_no_www}"

                safe = re.sub(r"[^\w]", "_", bare_no_www)
                safe = re.sub(r"_+", "_", safe).strip("_")
                safe = safe[:80] or "target"

                out_json = f'"$HOME/recon/ffuf-{safe}.json"'
                expected_jsons.append(f"ffuf-{safe}.json")

                if "Directory" in mode:
                    fuzz_url = f"{canonical}/FUZZ"
                elif "Subdomain" in mode:
                    fuzz_url = f"https://FUZZ.{bare_no_www}"
                elif "Parameter" in mode:
                    fuzz_url = f"{canonical}?FUZZ=watchtower"
                else:
                    fuzz_url       = canonical
                    extra_headers += f' -H "Host: FUZZ.{bare_no_www}"'

                cmd = " ".join(filter(None, [
                    "ffuf",
                    f"-u '{fuzz_url}'",
                    f"-w {wordlist_remote}",
                    f"-t {threads}",
                    f"-timeout {timeout}",
                    f"-fc {filter_codes}",
                    follow_redirects,
                    ext_flag,
                    extra_headers.strip(),
                    rate_flag,
                    "-noninteractive -s",   # -s silences screen, -o still writes JSON file
                    f"-o {out_json}",
                    "-of json",
                ]))

                sh.write(f"echo '[ffuf] [{node}] Starting: {url}'\n")
                sh.write(cmd + " || true\n")
                sh.write(f"echo '[ffuf] [{node}] Done: {url}'\n\n")

            sh.write("echo '[ffuf] All targets complete on this node.'\n")

        # Preview script
        info(f"Script for {node}:")
        with open(script_path) as f:
            for line in f:
                print(f"    {line}", end="")
        print()

    # ── Upload wordlist once (broadcast to all nodes) ─────────────────────
    info(f"Preparing fleet nodes …")
    run_cmd(f"axiom-exec 'mkdir -p \"$HOME/recon\"' --fleet '{fleet}'", verbose=verbose)

    if upload_wordlist:
        info(f"Uploading wordlist '{wl_name}' to all fleet nodes …")
        run_cmd(
            f"axiom-scp \"{wordlist_local}\" '{fleet}*:recon/{wl_name}'",
            verbose=verbose,
        )

    # ── Upload per-node scripts and execute in parallel via SSH ──────────
    # We upload each node's script individually then trigger them all at once
    # using axiom-exec (which runs on all nodes in parallel and blocks until done)
    #
    # Since all nodes got different script names (ffuf-targets-pratik01.sh etc.)
    # we use a small dispatcher: each node runs its own named script.

    info(f"Uploading per-node scripts …")
    for node, script_path in per_node_scripts.items():
        run_cmd(
            f"axiom-scp \"{script_path}\" '{node}:recon/{script_path.name}'",
            verbose=verbose,
        )

    # Run each node's own script — axiom-exec broadcasts the same command to all
    # nodes, but each node's script is named after itself so it only runs its targets
    info(f"Executing ffuf on all nodes in parallel (blocking until complete) …")
    run_cmd(
        f"axiom-exec 'bash \"$HOME/recon/ffuf-targets-$(hostname).sh\" 2>&1 || true' "
        f"--fleet '{fleet}'",
        verbose=verbose,
    )

    # ── Pull all JSON results back ────────────────────────────────────────
    # ── Pull results — exact node:file per target, no glob ────────────────
    # axiom-scp strips * from the remote path (uses it only for node matching),
    # so we pull each JSON by its exact known name from its exact known node.
    info("Pulling ffuf JSON results from fleet …")
    pull_ok: list[str] = []
    pull_fail: list[str] = []

    for node, targets in node_targets.items():
        for url in targets:
            url         = url.strip().rstrip("/")
            bare        = re.sub(r"^https?://", "", url, flags=re.I)
            bare_no_www = re.sub(r"^www\.", "", bare, flags=re.I)
            safe        = re.sub(r"[^\w]", "_", bare_no_www)
            safe        = re.sub(r"_+", "_", safe).strip("_")
            safe        = safe[:80] or "target"
            json_name   = f"ffuf-{safe}.json"

            result = run_cmd(
                f"axiom-scp '{node}:recon/{json_name}' \"{raw_out}/\"",
                check=False, verbose=verbose,
            )
            if result.returncode == 0:
                pull_ok.append(json_name)
            else:
                pull_fail.append(f"{node}:{json_name}")

    json_files = list(raw_out.glob("*.json"))
    if not json_files:
        warn("No ffuf JSON results pulled from fleet.")
        warn("Expected: " + ", ".join(expected_jsons))
        warn("SSH into a fleet node and check: ls -la ~/recon/ffuf-*.json")
    else:
        success(f"Pulled {len(json_files)} result file(s): {[f.name for f in json_files]}")
        if pull_fail:
            warn("Failed to pull: " + ", ".join(pull_fail))

    # ── Parse → final outputs ─────────────────────────────────────────────
    out_txt  = fin / f"ffuf-findings-{_SESSION['timestamp']}.txt"
    out_xlsx = fin / f"ffuf-findings-{_SESSION['timestamp']}.xlsx"

    _ffuf_json_to_txt(raw_out, out_txt)
    _ffuf_json_to_xlsx(raw_out, str(out_xlsx))

    write_manifest_entry(MODULE, str(out_txt),  "ffuf bruteforce findings — plain text")
    write_manifest_entry(MODULE, str(out_xlsx), "ffuf bruteforce findings — Excel")

    success(f"ffuf complete. Final outputs → {fin}/")


def _ffuf_json_to_txt(results_dir: Path, out_path: Path) -> None:
    import json, re as _re

    header = _write_text_header(
        "ffuf",
        "Directory / endpoint bruteforce findings via ffuf",
        ["Columns: Status | Lines | Words | Size | Redirect | URL"],
    )
    lines: list[str] = []
    total_found = 0

    for jf in sorted(results_dir.glob("*.json")):
        try:
            data    = json.loads(jf.read_text(errors="replace"))
            results = data.get("results", [])

            # Parse wordlist path out of the commandline string
            cmdline  = data.get("commandline", "")
            wl_match = _re.search(r"-w\s+(\S+)", cmdline)
            wordlist = wl_match.group(1) if wl_match else "unknown"

            # Parse target URL out of commandline
            url_match = _re.search(r"-u\s+'?(\S+?)'?\s", cmdline)
            target_url = url_match.group(1) if url_match else jf.stem

            found = len(results)
            total_found += found

            lines.append(f"\n# {'═' * 68}")
            lines.append(f"# Target  : {target_url}")
            lines.append(f"# Wordlist: {wordlist}")
            lines.append(f"# Found   : {found} result(s)")
            lines.append(f"# {'═' * 68}")

            if not results:
                lines.append("# (no results matched the filter/matcher criteria)")
                continue

            lines.append(f"{'Status':<8} {'Lines':<7} {'Words':<7} {'Size':<8} {'Redirect':<45} URL")
            lines.append(f"# {'─' * 68}")
            for r in sorted(results, key=lambda x: x.get("status", 0)):
                redirect = r.get("redirectlocation", "") or ""
                lines.append(
                    f"{r.get('status', '?'):<8} "
                    f"{r.get('lines', '?'):<7} "
                    f"{r.get('words', '?'):<7} "
                    f"{r.get('length', '?'):<8} "
                    f"{redirect:<45} "
                    f"{r.get('url', '')}"
                )
        except (json.JSONDecodeError, KeyError) as exc:
            lines.append(f"# ERROR reading {jf.name}: {exc}")

    lines.append(f"\n# {'═' * 68}")
    lines.append(f"# Total findings across all targets: {total_found}")
    lines.append(f"# {'═' * 68}")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(header)
        f.write("\n".join(lines) + "\n")


def _ffuf_json_to_xlsx(results_dir: Path, out_path: str) -> None:
    """Convert ffuf JSON results to a structured Excel workbook."""
    import json
    from openpyxl import Workbook

    wb = Workbook()
    st = _excel_styles()
    _xlsx_add_meta_sheet(wb, "ffuf", "Directory / endpoint bruteforce via ffuf")

    ws = wb.create_sheet("Findings")
    HDRS = ["Target", "URL", "Status Code", "Lines", "Words", "Size (bytes)", "Redirect URL"]
    _write_header_row(ws, HDRS, st)

    row_idx = 2
    for jf in sorted(results_dir.glob("*.json")):
        try:
            data    = json.loads(jf.read_text(errors="replace"))
            results = data.get("results", [])
            for r in results:
                status = r.get("status", 0)
                # Colour code by status family
                if status < 300:
                    fill = st["alt_fill"]        # blue-ish = 2xx
                elif status < 400:
                    fill = st["white_fill"]       # white    = 3xx redirect
                else:
                    fill = st["red_fill"]         # red      = 4xx/5xx

                vals = [
                    data.get("config", {}).get("url", jf.stem),
                    r.get("url", ""),
                    status,
                    r.get("lines", ""),
                    r.get("words", ""),
                    r.get("length", ""),
                    r.get("redirectlocation", ""),
                ]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=row_idx, column=col, value=val)
                    c.fill      = fill
                    c.alignment = st["top"]
                row_idx += 1
        except (json.JSONDecodeError, KeyError):
            pass

    _set_col_widths(ws, [30, 70, 12, 8, 8, 14, 50])
    _freeze_and_filter(ws)
    wb.save(out_path)


# ─────────────────────────────────────────────
# Dispatch table
# ─────────────────────────────────────────────
DISPATCH: dict[str, callable] = {
    "nmap":         lambda a: run_nmap(a.input, a.fleet, a.verbose),
    "getheaders":   lambda a: run_getheaders(a.input, a.verbose),
    "getbody":      lambda a: run_getbody(a.input, a.verbose),
    "spidering":    lambda a: run_spidering(a.input, a.fleet, a.verbose),
    "secretfinder": lambda a: run_secretfinder(a.input, a.fleet, a.verbose),
    "sslscan":      lambda a: run_sslscan(a.input, a.fleet, a.verbose),
    "subdomains":   lambda a: run_subdomains(a.input, a.fleet, a.verbose),
    "whatweb":      lambda a: run_whatweb(a.input, a.fleet, a.verbose),
    "waf":          lambda a: run_waf(a.input, a.fleet, a.verbose),
    "spiderfoot":   lambda a: run_spiderfoot(a.input, a.verbose),
    "formfinder":   lambda a: run_formfinder(a.input, a.verbose),
    "dnsrecon":     lambda a: run_dnsrecon(a.input, a.verbose),
    "waybackurls":  lambda a: run_waybackurls(a.input, a.fleet, a.verbose),
    "httpx":        lambda a: run_httpx(a.input, a.fleet, a.verbose),
    "nuclei":       lambda a: run_nuclei(a.input, a.fleet, a.verbose),
    "finddocs":     lambda a: run_finddocs(a.input, a.extension, a.verbose),
    "asn":          lambda a: run_asn(a.input, a.verbose),
    "cve":          lambda a: run_cve(a.input, a.verbose),
    "github":       lambda a: run_github(a.input, a.verbose),
    "gitleaks":     lambda a: run_gitleaks(a.input, a.verbose),
    "ffuf":         lambda a: run_ffuf(a.input, a.fleet, a.verbose),
}

# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────
def _epilog() -> str:
    lines = [bold("\nAvailable Modules:"), "", bold("  Requiring axiom-scan:")]
    for m in AXIOM_MODULES:
        lines.append(f"    {cyan(m):<22}  {MODULE_REGISTRY[m]['desc']}")
    lines += ["", bold("  Standalone:")]
    for m in MODULES:
        if m not in AXIOM_MODULES:
            lines.append(f"    {cyan(m):<22}  {MODULE_REGISTRY[m]['desc']}")
    lines += [
        "",
        bold("Output structure:"),
        "  watchtower-results/",
        "  ├── MANIFEST.txt          ← index of every output file",
        "  ├── final/<module>/       ← analyst deliverables (timestamped)",
        "  └── debug/<module>/       ← raw/intermediate artefacts",
        "",
        bold("Examples:"),
        f"  {green('watchtower.py')} -m nmap        -i targets.txt  -f my_fleet",
        f"  {green('watchtower.py')} -m spidering   -i domains.txt  -f osint",
        f"  {green('watchtower.py')} -m dnsrecon    -i domains.txt",
        f"  {green('watchtower.py')} -m getheaders  -i targets.txt",
        f"  {green('watchtower.py')} -m finddocs    -i urls.txt     -e pdf",
        "",
        bold("Notes:"),
        "  • axiom-scan modules require an active axiom fleet (-f).",
        "  • All final outputs are in watchtower-results/final/<module>/",
        "  • Every output file has a metadata header (who/what/when/how).",
        "  • Intermediate files go to watchtower-results/debug/<module>/",
        "  • Use NO_COLOR=1 to disable colour output.",
        "  • Use DEBUG=1 for full tracebacks on errors.",
        "  • Use -v / --verbose for detailed command output.",
        "",
    ]
    return "\n".join(lines)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="watchtower",
        description=f"{BANNER}\n{bold('Automated Reconnaissance & OSINT Framework')}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=_epilog(),
    )
    p.add_argument("-m", "--module",    metavar="MODULE",  required=True,
                   choices=MODULES,
                   help="Module to run (see list below)")
    p.add_argument("-i", "--input",     metavar="FILE",    required=True,
                   help="Input file (domains or URLs, one per line)")
    p.add_argument("-f", "--fleet",     metavar="FLEET",   default="",
                   help="Axiom fleet name (required for axiom-scan modules)")
    p.add_argument("-e", "--extension", metavar="EXT",     default="",
                   help="File extension for 'finddocs' module (e.g. pdf)")
    p.add_argument("-v", "--verbose",   action="store_true",
                   help="Show all commands being executed")
    p.add_argument("--version",         action="version",  version="WatchTower v1.0.0")
    return p


def main() -> None:
    print(BANNER)
    _setup_completion()
    parser = build_parser()
    args   = parser.parse_args()

    init_session(args)
    _init_manifest()

    info(f"Module  : {bold(args.module)}")
    info(f"Input   : {args.input}")
    info(f"Output  : {bold(str(_output_root() / 'final' / args.module))}")
    if args.fleet:
        info(f"Fleet   : {args.fleet}")
    if args.extension:
        info(f"Ext     : {args.extension}")
    if args.verbose:
        info("Verbose : enabled")
    print()

    try:
        DISPATCH[args.module](args)
    except KeyboardInterrupt:
        print()
        warn("Interrupted by user.")
        sys.exit(130)
    except SystemExit:
        raise
    except Exception as exc:
        error(f"Unexpected error in module '{args.module}': {exc}")
        if os.environ.get("DEBUG"):
            import traceback
            traceback.print_exc()
        else:
            info("Run with DEBUG=1 for a full traceback.")
        sys.exit(1)

    # Print manifest location as a final reminder
    manifest = _output_root() / "MANIFEST.txt"
    print()
    success(f"All outputs → {bold(str(_output_root() / 'final' / args.module))}/")
    info(f"Full file index → {manifest}")


if __name__ == "__main__":
    main()
